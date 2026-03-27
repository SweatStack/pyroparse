#!/usr/bin/env python3
"""Generate src/fit/profile.rs from FIT SDK Profile.xlsx.

Reads the Profile.xlsx spreadsheet and a TOML config to produce a Rust
source file with message constants, field definition tables, and enum
name-mapping functions.

Usage:
    python scripts/generate_profile.py \\
        --sdk assets/Profile.xlsx \\
        --config scripts/profile.toml
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

if sys.version_info >= (3, 11):
    import tomllib
else:
    try:
        import tomllib
    except ImportError:
        import tomli as tomllib  # type: ignore[no-redef]


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------


@dataclass
class EnumValue:
    name: str
    value: int
    comment: str


@dataclass
class EnumType:
    name: str
    base_type: str
    values: list[EnumValue] = field(default_factory=list)


@dataclass
class ComponentDef:
    dest_field_name: str
    bits: int
    accumulate: bool


@dataclass
class FieldDef:
    number: int
    name: str
    field_type: str  # base type name or enum type name
    scale: float
    offset: float
    units: str
    components: list[ComponentDef] = field(default_factory=list)
    accumulate: bool = False


@dataclass
class MessageDef:
    name: str
    fields: list[FieldDef] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Parse Profile.xlsx
# ---------------------------------------------------------------------------


def parse_types_sheet(ws) -> dict[str, EnumType]:
    """Parse the Types sheet into a dict of enum type name -> EnumType."""
    types: dict[str, EnumType] = {}
    current: EnumType | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        type_name, base_type, value_name, value, comment = (
            row[0],
            row[1],
            row[2],
            row[3],
            row[4] if len(row) > 4 else "",
        )

        if type_name:
            current = EnumType(name=str(type_name), base_type=str(base_type))
            types[current.name] = current
        elif current and value_name and value is not None:
            # Parse value: may be hex string like "0xFF00" or int.
            v = _parse_int(value)
            if v is not None:
                current.values.append(
                    EnumValue(
                        name=str(value_name),
                        value=v,
                        comment=str(comment or ""),
                    )
                )

    return types


def parse_messages_sheet(ws) -> dict[str, MessageDef]:
    """Parse the Messages sheet into a dict of message name -> MessageDef."""
    messages: dict[str, MessageDef] = {}
    current: MessageDef | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        msg_name = row[0]
        field_num = row[1]
        field_name = row[2]
        field_type = row[3]
        _array = row[4]
        components = row[5]
        scale = row[6]
        offset = row[7]
        units = row[8]
        bits = row[9]
        accumulate = row[10]

        # New message block.
        if msg_name:
            current = MessageDef(name=str(msg_name))
            messages[current.name] = current
            continue

        # Skip subfield rows (no field number) — they share the parent's
        # field number and only affect naming/enum resolution.
        if not current or field_num is None or str(field_num).strip() == "":
            continue

        comps = _parse_components(components, bits, accumulate)
        field_accum = _is_accumulated(accumulate, comps)

        current.fields.append(
            FieldDef(
                number=int(field_num),
                name=str(field_name or ""),
                field_type=str(field_type or ""),
                scale=_parse_first_float(scale, 1.0),
                offset=_parse_first_float(offset, 0.0),
                units=str(units or ""),
                components=comps,
                accumulate=field_accum,
            )
        )

    return messages


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def _parse_int(value) -> int | None:
    """Parse an integer from a cell value (may be int, float, or hex string)."""
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        if s.startswith("0x") or s.startswith("0X"):
            return int(s, 16)
        return int(s)
    except ValueError:
        return None


def _parse_components(
    components: str | None, bits: str | None, accumulate: str | None
) -> list[ComponentDef]:
    """Parse comma-separated Components/Bits/Accumulate columns."""
    if not components:
        return []

    comp_names = [c.strip() for c in str(components).split(",")]
    bit_values = [int(b.strip()) for b in str(bits).split(",")] if bits else []
    accum_values = (
        [v.strip() == "1" for v in str(accumulate).split(",")]
        if accumulate
        else []
    )

    result = []
    for i, name in enumerate(comp_names):
        if not name:
            continue
        result.append(
            ComponentDef(
                dest_field_name=name,
                bits=bit_values[i] if i < len(bit_values) else 0,
                accumulate=accum_values[i] if i < len(accum_values) else False,
            )
        )
    return result


def _parse_first_float(value, default: float) -> float:
    """Parse a float from a cell that may be comma-separated (e.g. '100,16').

    For fields with components, the scale/offset columns contain one value
    per component. The field's own scale/offset is the first value.
    """
    if not value:
        return default
    s = str(value).split(",")[0].strip()
    try:
        return float(s)
    except ValueError:
        return default


def _is_accumulated(accumulate: str | None, components: list[ComponentDef]) -> bool:
    """Check if the field itself (not its components) is accumulated."""
    if not accumulate:
        return False
    # If there are components, the accumulate column applies to them, not
    # the field itself.
    if components:
        return False
    return str(accumulate).strip() in ("1", "true", "True")


# ---------------------------------------------------------------------------
# Rust code generation
# ---------------------------------------------------------------------------

# Map FIT profile base type names to our Rust BaseType enum variants.
BASE_TYPE_MAP = {
    "enum": "Enum",
    "sint8": "SInt8",
    "uint8": "UInt8",
    "sint16": "SInt16",
    "uint16": "UInt16",
    "sint32": "SInt32",
    "uint32": "UInt32",
    "string": "String",
    "float32": "Float32",
    "float64": "Float64",
    "uint8z": "UInt8z",
    "uint16z": "UInt16z",
    "uint32z": "UInt32z",
    "byte": "Byte",
    "sint64": "SInt64",
    "uint64": "UInt64",
    "uint64z": "UInt64z",
    "bool": "Enum",  # bool is encoded as enum in FIT
}


def _rust_base_type(field_type: str, all_types: dict[str, EnumType]) -> str:
    """Resolve a field type to a Rust BaseType variant name.

    If the field type is an enum name from the Types sheet, look up its
    base type. Otherwise, it should be a base type name directly.
    """
    # Direct base type?
    if field_type in BASE_TYPE_MAP:
        return BASE_TYPE_MAP[field_type]

    # Enum type — look up its underlying base type.
    if field_type in all_types:
        underlying = all_types[field_type].base_type
        if underlying in BASE_TYPE_MAP:
            return BASE_TYPE_MAP[underlying]

    # Fallback for unknown types.
    return "UInt8"


def _rust_float(value: float) -> str:
    """Format a float for Rust — always includes a decimal point."""
    if value == int(value):
        return f"{int(value)}.0"
    return f"{value:g}"


def _rust_field_type(field_type: str, all_types: dict[str, EnumType]) -> str:
    """Return Rust FieldType expression for a profile field type."""
    if field_type in BASE_TYPE_MAP:
        return f"FieldType::Base(BaseType::{BASE_TYPE_MAP[field_type]})"
    if field_type in all_types:
        return f'FieldType::Enum("{field_type}")'
    return f"FieldType::Base(BaseType::UInt8)"


def _to_screaming_snake(name: str) -> str:
    return name.upper()


def generate_rust(
    config: dict,
    all_types: dict[str, EnumType],
    all_messages: dict[str, MessageDef],
) -> str:
    """Generate the full profile.rs source."""
    lines: list[str] = []
    w = lines.append

    sdk_version = config.get("sdk_version", "unknown")
    included_messages = config.get("messages", {}).get("include", [])
    included_enums = config.get("enums", {}).get("include", [])

    # --- Header ---
    w(f'//! Auto-generated FIT profile from SDK {sdk_version}.')
    w("//!")
    w("//! DO NOT EDIT — regenerate with:")
    w("//!   python scripts/generate_profile.py --sdk assets/Profile.xlsx --config scripts/profile.toml")
    w("")
    w("#![allow(dead_code)]")
    w("")

    # --- BaseType enum ---
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("// Base types")
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("")
    w("#[derive(Debug, Clone, Copy, PartialEq, Eq)]")
    w("#[repr(u8)]")
    w("pub enum BaseType {")
    for name, variant in BASE_TYPE_MAP.items():
        if name in ("bool",):
            continue
        w(f"    {variant},")
    w("}")
    w("")
    w("impl BaseType {")
    w("    /// Size in bytes for this base type.")
    w("    pub const fn size(self) -> usize {")
    w("        match self {")
    w("            Self::Enum | Self::SInt8 | Self::UInt8 | Self::UInt8z | Self::Byte => 1,")
    w("            Self::SInt16 | Self::UInt16 | Self::UInt16z => 2,")
    w("            Self::SInt32 | Self::UInt32 | Self::UInt32z | Self::Float32 => 4,")
    w("            Self::String => 1,")
    w("            Self::Float64 | Self::SInt64 | Self::UInt64 | Self::UInt64z => 8,")
    w("        }")
    w("    }")
    w("")
    w("    /// Parse from the base type field byte (bits 0-4).")
    w("    pub const fn from_byte(b: u8) -> Self {")
    w("        match b & 0x1F {")
    w("            0 => Self::Enum,")
    w("            1 => Self::SInt8,")
    w("            2 => Self::UInt8,")
    w("            3 => Self::SInt16,")
    w("            4 => Self::UInt16,")
    w("            5 => Self::SInt32,")
    w("            6 => Self::UInt32,")
    w("            7 => Self::String,")
    w("            8 => Self::Float32,")
    w("            9 => Self::Float64,")
    w("            10 => Self::UInt8z,")
    w("            11 => Self::UInt16z,")
    w("            12 => Self::UInt32z,")
    w("            13 => Self::Byte,")
    w("            14 => Self::SInt64,")
    w("            15 => Self::UInt64,")
    w("            16 => Self::UInt64z,")
    w("            _ => Self::Byte,  // unknown types treated as opaque bytes")
    w("        }")
    w("    }")
    w("}")
    w("")

    # --- FieldType enum ---
    w("#[derive(Debug, Clone, Copy, PartialEq)]")
    w("pub enum FieldType {")
    w("    Base(BaseType),")
    w("    Enum(&'static str),")
    w("}")
    w("")

    # --- ComponentDef ---
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("// Field and component definitions")
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("")
    w("#[derive(Debug, Clone)]")
    w("pub struct ComponentDef {")
    w("    pub dest_field_name: &'static str,")
    w("    pub bits: u8,")
    w("    pub accumulate: bool,")
    w("}")
    w("")

    # --- FieldDef ---
    w("#[derive(Debug, Clone)]")
    w("pub struct FieldDef {")
    w("    pub number: u8,")
    w("    pub name: &'static str,")
    w("    pub base_type: BaseType,")
    w("    pub field_type: FieldType,")
    w("    pub scale: f64,")
    w("    pub offset: f64,")
    w("    pub units: &'static str,")
    w("    pub components: &'static [ComponentDef],")
    w("    pub accumulate: bool,")
    w("}")
    w("")
    w("impl FieldDef {")
    w("    /// Look up a field by number in a field table.")
    w("    pub fn lookup(table: &[FieldDef], number: u8) -> Option<&FieldDef> {")
    w("        table.iter().find(|f| f.number == number)")
    w("    }")
    w("}")
    w("")

    # --- Message number constants ---
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("// Message numbers")
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("")

    # We need the mesg_num enum to resolve message names → numbers.
    mesg_num_type = all_types.get("mesg_num")
    mesg_num_map: dict[str, int] = {}
    if mesg_num_type:
        mesg_num_map = {v.name: v.value for v in mesg_num_type.values}

    for msg_name in included_messages:
        num = mesg_num_map.get(msg_name)
        if num is not None:
            w(f"pub const MESG_{_to_screaming_snake(msg_name)}: u16 = {num};")
        else:
            w(f"// WARNING: message '{msg_name}' not found in mesg_num type")

    w("")

    # --- FIT epoch ---
    w("/// Seconds between Unix epoch (1970-01-01) and FIT epoch (1989-12-31).")
    w("pub const FIT_EPOCH_OFFSET: i64 = 631_065_600;")
    w("")
    w("/// Timestamps below this threshold are relative, not absolute FIT epoch values.")
    w("pub const DATETIME_MIN: u32 = 0x10000000;")
    w("")

    # --- Field tables per message ---
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("// Field definitions per message")
    w("// ═══════════════════════════════════════════════════════════════════════════")

    for msg_name in included_messages:
        msg = all_messages.get(msg_name)
        if msg is None:
            w(f"\n// WARNING: message '{msg_name}' not found in Messages sheet")
            continue

        w("")
        w(f"pub const {_to_screaming_snake(msg_name)}_FIELDS: &[FieldDef] = &[")

        for f in msg.fields:
            base = _rust_base_type(f.field_type, all_types)
            ftype = _rust_field_type(f.field_type, all_types)

            # Component definitions.
            if f.components:
                comp_parts = []
                for c in f.components:
                    comp_parts.append(
                        f'ComponentDef {{ dest_field_name: "{c.dest_field_name}", '
                        f"bits: {c.bits}, accumulate: {'true' if c.accumulate else 'false'} }}"
                    )
                comp_expr = f"&[{', '.join(comp_parts)}]"
            else:
                comp_expr = "&[]"

            w(f"    FieldDef {{")
            w(f"        number: {f.number},")
            w(f'        name: "{f.name}",')
            w(f"        base_type: BaseType::{base},")
            w(f"        field_type: {ftype},")
            w(f"        scale: {_rust_float(f.scale)},")
            w(f"        offset: {_rust_float(f.offset)},")
            w(f'        units: "{f.units}",')
            w(f"        components: {comp_expr},")
            w(f"        accumulate: {'true' if f.accumulate else 'false'},")
            w(f"    }},")

        w("];")

    w("")

    # --- Enum name functions ---
    w("// ═══════════════════════════════════════════════════════════════════════════")
    w("// Enum name mappings")
    w("// ═══════════════════════════════════════════════════════════════════════════")

    for enum_name in included_enums:
        enum_type = all_types.get(enum_name)
        if enum_type is None:
            w(f"\n// WARNING: enum type '{enum_name}' not found in Types sheet")
            continue

        # Determine the Rust parameter type from the enum's base type.
        rust_param = _rust_param_type(enum_type.base_type)

        w("")
        w(f"/// Map {enum_name} numeric value to its string name.")
        w(f"pub fn {enum_name}_name(val: {rust_param}) -> &'static str {{")
        w(f"    match val {{")

        for v in enum_type.values:
            w(f'        {v.value} => "{v.name}",')

        w(f'        _ => "unknown",')
        w(f"    }}")
        w(f"}}")

    w("")
    return "\n".join(lines)


def _rust_param_type(base_type: str) -> str:
    """Map a FIT base type to a Rust integer parameter type for match arms."""
    return {
        "enum": "u8",
        "uint8": "u8",
        "uint8z": "u8",
        "sint8": "i8",
        "uint16": "u16",
        "uint16z": "u16",
        "sint16": "i16",
        "uint32": "u32",
        "uint32z": "u32",
        "sint32": "i32",
        "uint64": "u64",
        "uint64z": "u64",
        "sint64": "i64",
    }.get(base_type, "u16")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate src/fit/profile.rs from FIT SDK Profile.xlsx"
    )
    parser.add_argument(
        "--sdk",
        type=Path,
        default=Path("assets/Profile.xlsx"),
        help="Path to Profile.xlsx (default: assets/Profile.xlsx)",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("scripts/profile.toml"),
        help="Path to profile.toml config (default: scripts/profile.toml)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("src/fit/profile.rs"),
        help="Output path (default: src/fit/profile.rs)",
    )
    args = parser.parse_args()

    # Load config.
    with open(args.config, "rb") as f:
        config = tomllib.load(f)

    # Load Profile.xlsx.
    wb = openpyxl.load_workbook(args.sdk, read_only=True)
    all_types = parse_types_sheet(wb["Types"])
    all_messages = parse_messages_sheet(wb["Messages"])

    print(f"Parsed {len(all_types)} types, {len(all_messages)} messages from {args.sdk}")

    # Generate.
    rust_source = generate_rust(config, all_types, all_messages)

    # Write.
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(rust_source)

    lines = rust_source.count("\n") + 1
    print(f"Generated {args.output} ({lines} lines)")


if __name__ == "__main__":
    main()
